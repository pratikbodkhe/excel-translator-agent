import openpyxl
import hashlib
import shutil
import json
import re
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
import logging

from src.cache_manager import MultiLayerCache, RedisCache, PostgreSQLCache, NoCache
from src.llm_providers import (
    BaseLLMProvider,
    TranslationRequest,
    BatchTranslationRequest,
    BatchTranslationResponse,
    create_llm_provider,
)
from src.config.config import config

logger = logging.getLogger(__name__)


@dataclass
class CellData:
    sheet_name: str
    cell_ref: str
    content: str
    context: str = ""  # e.g., "header", "footer", etc.


class LLMBatchProcessor:
    def __init__(
        self,
        llm_provider: BaseLLMProvider,
        max_batch_size: int = 50,
        additional_context: str = "",
    ):
        self.llm_provider = llm_provider
        self.max_batch_size = max_batch_size
        self.additional_context = additional_context

    def translate_batch(self, batch: List[CellData]) -> Dict[str, str]:
        # Prepare data for LLM using the new format
        translation_requests = [
            TranslationRequest(
                id=f"{cell.sheet_name}!{cell.cell_ref}",
                text=cell.content,
                context=cell.context,
            )
            for cell in batch
        ]

        batch_request = BatchTranslationRequest(
            translations=translation_requests,
            batch_id=hashlib.md5(
                json.dumps([c.__dict__ for c in batch]).encode()
            ).hexdigest(),
            metadata={
                "file_hash": hashlib.md5(
                    json.dumps([c.__dict__ for c in batch]).encode()
                ).hexdigest(),
                "sheet_name": batch[0].sheet_name if batch else "",
            },
            additional_context=self.additional_context,
        )

        # Send to LLM and get response
        response = self.llm_provider.translate_batch(batch_request)

        # Convert response to dict mapping cell IDs to translations
        return {t.id: t.text for t in response.translations}


class ExcelTranslator:
    def __init__(
        self, llm_provider, cache, batch_size: int = 50, additional_context: str = ""
    ):
        self.llm_provider = llm_provider
        self.cache = cache
        self.batch_size = batch_size
        self.additional_context = additional_context
        self.batch_processor = LLMBatchProcessor(
            llm_provider, batch_size, additional_context
        )

    def _should_translate_cell(self, cell) -> bool:
        # Skip empty cells, numbers, formulas, dates, years, and special characters
        if cell is None or cell.value is None:
            return False

        # Skip numeric values
        if isinstance(cell.value, (int, float)):
            return False

        # Skip formulas
        if isinstance(cell.value, str) and cell.value.startswith("="):
            return False

        # Skip if the content is a string
        if isinstance(cell.value, str):
            value = str(cell.value).strip()

            # Skip if empty after stripping
            if not value:
                return False

            # Skip dates in common formats (YYYY-MM-DD, MM/DD/YYYY, etc.)
            if (
                re.match(r"^\d{4}-\d{1,2}-\d{1,2}", value)
                or re.match(r"^\d{1,2}/\d{1,2}/\d{4}", value)
                or re.match(r"^\d{1,2}-\d{1,2}-\d{4}", value)
                or re.match(r"^\d{2}\.\d{2}\.\d{4}", value)
            ):
                return False

            # Skip timestamps
            if re.match(r"^\d{4}-\d{1,2}-\d{1,2}[T\s]\d{1,2}:\d{1,2}", value):
                return False

            # Skip years (4 digit numbers)
            if re.match(r"^\d{4}$", value):
                return False

            # Skip special characters and symbols only content
            # This includes various Unicode symbols, punctuation, etc.
            if re.match(r"^[\W\d]+$", value) and not re.search(r"[a-zA-Z]", value):
                return False

            # Skip content that is just numbers and punctuation
            if re.match(r"^[\d\s\.,;:\-\+\*\/\\\(\)\[\]\{\}]+$", value):
                return False

        return True

    def _get_cell_context(self, sheet, cell) -> str:
        # Determine if cell is a header, footer, or regular cell
        if cell.row == 1:
            return "header"
        elif cell.row == sheet.max_row:
            return "footer"
        else:
            return "body"

    def _get_batch_for_sheet(
        self, sheet, sheet_name_for_id: str
    ) -> List[List[CellData]]:
        """Group cells into batches for translation."""
        batch = []
        batches = []

        for row in sheet.iter_rows():
            for cell in row:
                if self._should_translate_cell(cell):
                    context = self._get_cell_context(sheet, cell)
                    cell_data = CellData(
                        sheet_name=sheet_name_for_id,
                        cell_ref=cell.coordinate,
                        content=str(cell.value),
                        context=context,
                    )
                    batch.append(cell_data)

                    if len(batch) >= self.batch_size:
                        batches.append(batch)
                        batch = []

        if batch:  # Add any remaining cells as the last batch
            batches.append(batch)

        return batches

    def _translate_batch(self, batch: List[CellData]) -> Dict[str, str]:
        """Translate a batch of cells."""
        logger.debug(f"Translating batch of {len(batch)} cells")

        # Check cache first
        translations = {}
        to_translate = []

        for cell in batch:
            cached = self.cache.get_translation(cell.content, cell.context)
            if cached:
                logger.debug(f"Cache hit for '{cell.content}': {cached}")
                translations[f"{cell.sheet_name}!{cell.cell_ref}"] = cached
            else:
                logger.debug(f"Cache miss for '{cell.content}', adding to LLM batch")
                to_translate.append(cell)

        if to_translate:
            logger.debug(f"Sending {len(to_translate)} cells to LLM")
            # Process with LLM
            llm_translations = self.batch_processor.translate_batch(to_translate)
            logger.debug(f"LLM returned {len(llm_translations)} translations")

            for cell_id, translation in llm_translations.items():
                logger.debug(f"LLM translation: {cell_id} -> {translation}")

            translations.update(llm_translations)

            # Store in cache
            for cell in to_translate:
                cell_id = f"{cell.sheet_name}!{cell.cell_ref}"
                if cell_id in llm_translations:
                    self.cache.store_translation(
                        cell.content, cell.context, llm_translations[cell_id]
                    )

        return translations

    def _translate_sheet_name(self, sheet_name: str) -> str:
        """Translate a sheet name."""
        # Check cache first
        cached = self.cache.get_translation(sheet_name, "sheet_name")
        if cached:
            logger.debug(f"Cache hit for sheet name '{sheet_name}': {cached}")
            return cached

        # Create a single cell data for sheet name translation
        cell_data = CellData(
            sheet_name="",  # Empty for sheet names
            cell_ref="sheet_name",
            content=sheet_name,
            context="sheet_name",
        )

        # Translate using batch processor
        translations = self.batch_processor.translate_batch([cell_data])
        translated_name = translations.get("!sheet_name", sheet_name)

        # Store in cache
        self.cache.store_translation(sheet_name, "sheet_name", translated_name)

        logger.debug(f"Translated sheet name '{sheet_name}' to '{translated_name}'")
        return translated_name

    def translate_excel(self, input_path: str, output_path: str):
        """Main translation method that handles the entire process."""
        # Create a direct copy of the file to preserve all formatting, charts, etc.
        try:
            shutil.copy(input_path, output_path)
        except Exception as e:
            logger.error(f"Failed to create a copy of the workbook: {e}")
            return

        # Load the copied workbook for in-place translation
        wb = openpyxl.load_workbook(output_path)

        # First, translate all sheet names and rename them in the workbook
        original_sheet_names = wb.sheetnames
        sheet_name_mapping = {
            name: self._translate_sheet_name(name) for name in original_sheet_names
        }

        for original_name, translated_name in sheet_name_mapping.items():
            if original_name in wb:
                wb[original_name].title = translated_name
                logger.debug(f"Renamed sheet '{original_name}' to '{translated_name}'")

        # Now, process each sheet for cell translation
        for original_name, translated_name in sheet_name_mapping.items():
            logger.info(f"Processing sheet: {original_name} (now {translated_name})")
            sheet = wb[translated_name]  # Work with the sheet under its new name

            # Get batches, using the original name for consistent IDs
            batches = self._get_batch_for_sheet(sheet, original_name)
            for batch in batches:
                translations = self._translate_batch(batch)
                for cell_data in batch:
                    cell_id = f"{cell_data.sheet_name}!{cell_data.cell_ref}"
                    if cell_id in translations:
                        sheet[cell_data.cell_ref].value = translations[cell_id]

        # Save the modified workbook
        wb.save(output_path)
        logger.info(f"Translation complete. Saved to {output_path}")


class MCPAdapter:
    """Adapts the translator to work as an MCP server."""

    def __init__(self, translator: ExcelTranslator):
        self.translator = translator

    def handle_translation_request(self, input_path: str, output_path: str):
        try:
            self.translator.translate_excel(input_path, output_path)
            return {"status": "success", "message": "Translation complete"}
        except Exception as e:
            return {"status": "error", "message": str(e)}
