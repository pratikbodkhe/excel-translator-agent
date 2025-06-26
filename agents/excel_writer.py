import openpyxl
from openpyxl.styles import Alignment, Border, Fill, Font

from config.config import config


class ExcelWriter:
    def __init__(self, template_path):
        self.template_path = template_path
        self.workbook = openpyxl.load_workbook(template_path)

    def write_translated_sheet(self, sheet_name, translated_data):
        """Write translated data to a sheet while preserving formatting"""
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)

        sheet = self.workbook[sheet_name]
        sheet.delete_rows(1, sheet.max_row)  # Clear existing content

        for row_idx, row_data in enumerate(translated_data, 1):
            for col_idx, cell_data in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = cell_data['value']

                # Skip formatting for now to avoid parsing issues
                # TODO: Implement proper formatting preservation

    def _parse_style(self, style_str):
        """Parse style string into dictionary (simplified)"""
        # In real implementation, this would properly parse the style string
        # For now, return an empty dict as placeholder
        return {}

    def save(self, output_path):
        """Save the modified workbook"""
        self.workbook.save(output_path)
        self.workbook.close()
