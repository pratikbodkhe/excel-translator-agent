import openpyxl

from config import config


class ExcelReader:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.sheet_names = []

    def load_workbook(self):
        """Load the Excel workbook and get sheet names"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.sheet_names = self.workbook.sheetnames
            return True
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False

    def get_sheet_data(self, sheet_name):
        """Extract data from a specific sheet while preserving formatting"""
        if not self.workbook or sheet_name not in self.sheet_names:
            return None

        sheet = self.workbook[sheet_name]
        data = []

        # Iterate through rows and columns to extract cell data and formatting
        for row in sheet.iter_rows():
            row_data = []
            for cell in row:
                # Preserve cell value and formatting information
                cell_info = {
                    'value': cell.value,
                    'font': str(cell.font),
                    'fill': str(cell.fill),
                    'border': str(cell.border),
                    'alignment': str(cell.alignment),
                    'number_format': cell.number_format
                }
                row_data.append(cell_info)
            data.append(row_data)

        return data

    def close(self):
        """Close the workbook when done"""
        if self.workbook:
            self.workbook.close()
